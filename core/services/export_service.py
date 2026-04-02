"""
HRTech - Serviço de Exportação
==============================

Centraliza geração de planilhas Excel para exportações do RH.

SECURITY: Implementa mascaramento de dados sensíveis (LGPD).
"""

import io
import csv
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExportService:
    """Serviço de exportação com suporte a mascaramento de PII."""
    
    class _Echo:
        """Objeto compatível com csv.writer para streaming."""

        def write(self, value):
            return value

    @staticmethod
    def _mask_email(email: str) -> str:
        """
        Mascara email mantendo apenas parte do domínio visível.
        
        Exemplo: joao.silva@empresa.com → j***@empresa.com
        """
        if not email or '@' not in email:
            return email or '-'
        local, domain = email.rsplit('@', 1)
        if len(local) <= 1:
            masked_local = local[0] + '***'
        else:
            masked_local = local[0] + '***'
        return f"{masked_local}@{domain}"
    
    @staticmethod
    def _mask_phone(phone: str) -> str:
        """
        Mascara telefone mantendo apenas últimos 4 dígitos.
        
        Exemplo: (11) 98765-4321 → ***-4321
        """
        if not phone:
            return '-'
        # Remove caracteres não numéricos
        digits = ''.join(filter(str.isdigit, phone))
        if len(digits) < 4:
            return '***'
        return f"***-{digits[-4:]}"
    
    @staticmethod
    def _mask_name(name: str) -> str:
        """
        Mascara nome mantendo apenas iniciais.
        
        Exemplo: João da Silva → J. S.
        """
        if not name:
            return '-'
        parts = name.strip().split()
        if len(parts) == 1:
            return parts[0][0] + '.' if parts[0] else '-'
        initials = [p[0].upper() + '.' for p in parts if p]
        return ' '.join(initials)

    @staticmethod
    def _base_styles():
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        return header_font, header_fill, header_alignment, thin_border

    @classmethod
    def build_candidatos_workbook(cls, candidatos, mask_pii: bool = False):
        """
        Gera workbook Excel com lista de candidatos.
        
        Args:
            candidatos: QuerySet de candidatos
            mask_pii: Se True, mascara dados sensíveis (LGPD compliance)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidatos"

        header_font, header_fill, header_alignment, thin_border = cls._base_styles()

        # SECURITY: Headers não expõem dados sensíveis diretamente
        headers = ['Nome', 'Email', 'Telefone', 'Senioridade', 'Anos Exp.', 'Etapa', 'Status CV', 'Disponível', 'Cadastro']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row, candidato in enumerate(candidatos, 2):
            # SECURITY: Mascarar PII se solicitado
            if mask_pii:
                nome = cls._mask_name(candidato.nome)
                email = cls._mask_email(candidato.email)
                telefone = cls._mask_phone(candidato.telefone)
            else:
                nome = candidato.nome
                email = candidato.email
                telefone = candidato.telefone or '-'
            
            ws.cell(row=row, column=1, value=nome).border = thin_border
            ws.cell(row=row, column=2, value=email).border = thin_border
            ws.cell(row=row, column=3, value=telefone).border = thin_border
            ws.cell(row=row, column=4, value=candidato.get_senioridade_display()).border = thin_border
            ws.cell(row=row, column=5, value=candidato.anos_experiencia).border = thin_border
            ws.cell(row=row, column=6, value=candidato.get_etapa_processo_display()).border = thin_border
            ws.cell(row=row, column=7, value=candidato.get_status_cv_display()).border = thin_border
            ws.cell(row=row, column=8, value='Sim' if candidato.disponivel else 'Não').border = thin_border
            ws.cell(row=row, column=9, value=candidato.created_at.strftime('%d/%m/%Y')).border = thin_border

        column_widths = [30, 35, 15, 12, 10, 20, 18, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"candidatos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return output.read(), filename

    @classmethod
    def build_ranking_workbook(cls, vaga, matches, mask_pii: bool = False):
        """
        Gera workbook Excel com ranking de candidatos para uma vaga.
        
        Args:
            vaga: Objeto Vaga
            matches: QuerySet de AuditoriaMatch
            mask_pii: Se True, mascara dados sensíveis (LGPD compliance)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Ranking"

        header_font, header_fill, _, thin_border = cls._base_styles()

        ws.cell(row=1, column=1, value=f"Ranking - {vaga.titulo}").font = Font(bold=True, size=14)
        ws.merge_cells('A1:G1')
        ws.cell(row=2, column=1, value=f"Área: {vaga.area} | Senioridade: {vaga.get_senioridade_desejada_display()}")
        ws.merge_cells('A2:G2')

        headers = ['Posição', 'Nome', 'Email', 'Score', 'Senioridade', 'Etapa', 'Data Match']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        row_num = 5
        posicao = 1
        for match in matches:
            if not match.candidato:
                continue

            # SECURITY: Mascarar PII se solicitado
            if mask_pii:
                nome = cls._mask_name(match.candidato.nome)
                email = cls._mask_email(match.candidato.email)
            else:
                nome = match.candidato.nome
                email = match.candidato.email

            ws.cell(row=row_num, column=1, value=posicao).border = thin_border
            ws.cell(row=row_num, column=2, value=nome).border = thin_border
            ws.cell(row=row_num, column=3, value=email).border = thin_border
            cell_score = ws.cell(row=row_num, column=4, value=f"{match.score:.1f}%")
            cell_score.border = thin_border
            if match.score >= 80:
                cell_score.fill = PatternFill(start_color="d1e7dd", end_color="d1e7dd", fill_type="solid")
            elif match.score >= 60:
                cell_score.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            ws.cell(row=row_num, column=5, value=match.candidato.get_senioridade_display()).border = thin_border
            ws.cell(row=row_num, column=6, value=match.candidato.get_etapa_processo_display()).border = thin_border
            ws.cell(row=row_num, column=7, value=match.created_at.strftime('%d/%m/%Y %H:%M')).border = thin_border

            row_num += 1
            posicao += 1

        column_widths = [10, 30, 35, 12, 12, 18, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"ranking_{vaga.id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return output.read(), filename

    @classmethod
    def stream_candidatos_csv(cls, candidatos_iterator, mask_pii: bool = False):
        """
        Stream CSV de candidatos para download.
        
        Args:
            candidatos_iterator: Iterator de candidatos
            mask_pii: Se True, mascara dados sensíveis (LGPD compliance)
        """
        pseudo_buffer = cls._Echo()
        writer = csv.writer(pseudo_buffer)

        yield writer.writerow([
            'Nome', 'Email', 'Telefone', 'Senioridade', 'Anos Exp.',
            'Etapa', 'Status CV', 'Disponivel', 'Cadastro'
        ])

        for candidato in candidatos_iterator:
            # SECURITY: Mascarar PII se solicitado
            if mask_pii:
                nome = cls._mask_name(candidato.nome)
                email = cls._mask_email(candidato.email)
                telefone = cls._mask_phone(candidato.telefone)
            else:
                nome = candidato.nome
                email = candidato.email
                telefone = candidato.telefone or '-'
            
            yield writer.writerow([
                nome,
                email,
                telefone,
                candidato.get_senioridade_display(),
                candidato.anos_experiencia,
                candidato.get_etapa_processo_display(),
                candidato.get_status_cv_display(),
                'Sim' if candidato.disponivel else 'Nao',
                candidato.created_at.strftime('%d/%m/%Y'),
            ])
